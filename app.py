from __future__ import annotations

import asyncio
import base64
import csv
import dataclasses
import io
import json
import sqlite3
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import AsyncIterator, Optional

import httpx
import webview
from nicegui import app, run, ui
from openpyxl import load_workbook

# ui.chat_message's `avatar` expects an image URL, not raw text - a bare "🦉"
# renders as a broken-image placeholder. Encode it as an inline SVG data URI instead.
_OWL_SVG = (
    '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100">'
    '<text x="50" y="78" font-size="72" text-anchor="middle">🦉</text></svg>'
)
OWL_AVATAR = "data:image/svg+xml;base64," + base64.b64encode(_OWL_SVG.encode("utf-8")).decode("ascii")

# Vendored Tabler Icons (MIT) for the file explorer's per-type icons, kept
# fully offline (no CDN). These are consumed via a custom `default-header`
# slot on ui.tree (see render_explorer), not QIcon's `name` prop - Quasar's
# QIcon has a hardcoded prefix map where "ti-" already means Themify Icons,
# not Tabler, so a raw data URI rendered through our own <img> template
# sidesteps that collision entirely instead of fighting it.
_TABLER_ICONS_DIR = Path(__file__).resolve().parent / "assets" / "tabler_icons"


def _load_tabler_svg(slug: str, color: str) -> str:
    raw = (_TABLER_ICONS_DIR / f"{slug}.svg").read_text(encoding="utf-8")
    colored = raw.replace("currentColor", color)
    return "data:image/svg+xml;base64," + base64.b64encode(colored.encode("utf-8")).decode("ascii")


_FILETYPE_ICONS: dict[str, tuple[str, str]] = {
    "folder": ("folder", "#e8a838"),
    "js": ("file-type-js", "#f0db4f"),
    "jsx": ("file-type-jsx", "#61dafb"),
    "css": ("file-type-css", "#264de4"),
    "json": ("braces", "#8bc34a"),
    "md": ("markdown", "#8a8578"),
    "image": ("photo", "#e91e63"),
    "design": ("vector-bezier", "#a259ff"),
    "sheet": ("table", "#1d6f42"),
    "cfg": ("settings", "#8a8578"),
    "pdf": ("file-type-pdf", "#e53935"),
    "code": ("file-code", "#2dd4bf"),
    "default": ("file", "#8a8578"),
}
_TABLER_ICON_CACHE: dict[str, str] = {
    category: _load_tabler_svg(slug, color) for category, (slug, color) in _FILETYPE_ICONS.items()
}


def _tabler_icon(category: str) -> str:
    return _TABLER_ICON_CACHE[category]


from config import AGENTS_DIR, CONFIG, SKILLS_DIR, add_whitelisted_folder, remove_whitelisted_folder
from core import google_auth, location as location_service, memory, projects
from core.agents import Agent, load_agents
from core.llm import OllamaClient
from core.router import route_agent
from core.secrets import delete_secret, get_secret, set_secret
from core.skills import Skill, load_skills, match_skills
from core.tool_loop import Continue, FinalAnswer, PendingConfirmation, execute_tool, step_agent_stream
from core.tools import maps as maps_tools
from core.tools import websearch_tools
from core.tools.weather import get_weather
from core.tools.datetime_tools import get_current_time
from core.tools.filesystem import (
    create_folder, get_active_roots, read_file, read_file_bytes, set_active_roots, write_file,
)
from core.tools.memory_tools import set_memory_scope

_LANG_BY_EXT = {
    ".py": "python", ".js": "javascript", ".ts": "typescript", ".json": "json",
    ".yaml": "yaml", ".yml": "yaml", ".html": "html", ".css": "css",
    ".sh": "bash", ".sql": "sql", ".java": "java", ".go": "go", ".rs": "rust",
    ".c": "c", ".cpp": "cpp", ".rb": "ruby", ".php": "php", ".swift": "swift",
    ".toml": "toml", ".xml": "xml",
}

_MIME_BY_EXT = {
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp", ".svg": "image/svg+xml",
    ".pdf": "application/pdf",
}
_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}
_PDF_EXTS = {".pdf"}
_EXCEL_EXTS = {".xlsx", ".xlsm"}


# mistral has a strong training bias toward claiming it "can't know the real
# time", which fights the get_current_time tool call even when instructed
# (empirically ~15-30% success). Originally this was only injected when the
# user's message matched a narrow set of "what time is it"-style keywords,
# but that missed the much more common implicit case (scheduling requests
# like "creá un evento hoy...", "invitá a X el viernes") where the model
# still needs to know "now" but never says so explicitly - it silently
# defaulted to a training-era date (2023) instead. Unconditional and cheap,
# so now it's just always included.
def _time_context() -> str:
    info = get_current_time()
    return (
        f"The user's local system time right now is {info['time']} on {info['date']} "
        f"({info['weekday']}). Use this as the real current date/time for anything "
        f"relative ('today', 'tomorrow', 'this week') - never assume or guess a date."
    )


# Ported from a sibling project (AgenteCode, a separate local coding-agent
# CLI on the same qwen3:8b) where both lines were validated against
# concrete failures: the model fabricating a full "ran the script" output
# it never actually produced, and re-issuing the identical failing tool
# call instead of adjusting. Shared across all agents here (not duplicated
# per agents/*.yaml) so it can't drift out of sync between them.
def _location_context() -> str:
    loc = location_service.get_cached_location()
    if not loc:
        return ""
    place = ", ".join(p for p in (loc["city"], loc["region"], loc["country"]) if p)
    precision = (
        "precise, via the device's GPS/WiFi location"
        if loc.get("source") == "gps"
        else "approximate, via IP geolocation - may be off, especially with VPN"
    )
    return (
        f"The user's device appears to be located in {place} ({precision}). Use this only "
        "when the user refers to their current location ('acá', 'cerca mío', 'mi ubicación') "
        "without naming a place explicitly - otherwise trust what the user says."
    )


_WEEKDAYS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
_MONTHS_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def _format_header_date() -> str:
    now = datetime.now()
    return f"{_WEEKDAYS_ES[now.weekday()]} {now.day}, {_MONTHS_ES[now.month - 1]}"


# Process-wide, not per-session, same reasoning as location.py's cache - this is a
# single native window, not a multi-client web app.
_header_weather: dict = {}


_AGENT_GUARDRAILS = (
    "If a tool call fails or returns an error, read the message and adjust your "
    "next call - never repeat the exact same failing call again. Never claim you "
    "performed an action you did not actually call a tool for, and never invent "
    "a tool's output."
)


def _guess_language(ext: str) -> str | None:
    return _LANG_BY_EXT.get(ext)


# --- shared singletons (read-only infra, not per-conversation state) ---------

_client: OllamaClient | None = None


def get_client() -> OllamaClient:
    global _client
    if _client is None:
        _client = OllamaClient(CONFIG.ollama_base_url)
    return _client


def get_agents() -> dict[str, Agent]:
    return load_agents(AGENTS_DIR)


# --- state -------------------------------------------------------------------

@dataclass
class AppState:
    agent_id: str
    active_model: str
    active_project_id: int | None = None
    conversation_id: int | None = None
    messages: list[dict] = field(default_factory=list)
    skills: list[Skill] = field(default_factory=list)
    open_docs: list[dict] = field(default_factory=list)
    active_doc_path: str | None = None
    title_set: bool = False
    current_turn_task: Optional[asyncio.Task] = None


@dataclass
class UIRefs:
    chat_container: ui.column
    files_container: ui.column
    canvas_container: ui.column
    input_box: ui.input
    send_btn: ui.button
    content_splitter: ui.splitter


def _effective_agent(state: AppState) -> Agent:
    agent = get_agents()[state.agent_id]
    return dataclasses.replace(agent, model=state.active_model)


def _active_project_row(state: AppState) -> sqlite3.Row | None:
    if state.active_project_id is None:
        return None
    return memory.get_project(state.active_project_id)


# --- collapsible panels --------------------------------------------------------

_CANVAS_EXPANDED, _CANVAS_COLLAPSED = 62, 100


def _toggle_canvas(refs: UIRefs) -> None:
    s = refs.content_splitter
    s.value = _CANVAS_COLLAPSED if s.value < _CANVAS_COLLAPSED - 1 else _CANVAS_EXPANDED


# --- folder-backed projects ---------------------------------------------------

def open_folder_as_project(path: str) -> int:
    """Links an existing .lechu-marked folder to its project, or registers a new one."""
    folder = Path(path).expanduser().resolve(strict=False)
    marker = projects.read_marker(folder)
    if marker:
        existing = memory.get_project(marker.get("lechu_project_id", -1))
        if existing:
            return existing["id"]
    by_folder = memory.find_project_by_folder(str(folder))
    if by_folder:
        projects.write_marker(folder, by_folder["name"], by_folder["id"])
        return by_folder["id"]

    name = folder.name or str(folder)
    candidate = name
    suffix = 2
    while True:
        try:
            project_id = memory.create_project(candidate, folder_path=str(folder))
            break
        except sqlite3.IntegrityError:
            candidate = f"{name} ({suffix})"
            suffix += 1
    projects.write_marker(folder, candidate, project_id)
    return project_id


def sync_project_scope(project_row: sqlite3.Row | None) -> None:
    folder_path = project_row["folder_path"] if project_row else None
    set_active_roots(folder_path)
    set_memory_scope(folder_path)


async def _apply_active_project_scope(state: AppState) -> None:
    await run.io_bound(sync_project_scope, _active_project_row(state))


# --- conversation lifecycle -------------------------------------------------

async def start_new_conversation(state: AppState, agent: Agent) -> None:
    conv_id = await run.io_bound(
        memory.create_conversation, agent_id=agent.id, model=agent.model, project_id=state.active_project_id
    )
    await run.io_bound(memory.add_message, conv_id, "system", agent.system_prompt)
    state.conversation_id = conv_id
    state.messages = [{"role": "system", "content": agent.system_prompt}]
    state.title_set = False
    state.open_docs = []
    state.active_doc_path = None


async def load_conversation(state: AppState, conv_row: sqlite3.Row, agents: dict[str, Agent]) -> None:
    state.conversation_id = conv_row["id"]
    state.agent_id = conv_row["agent_id"] if conv_row["agent_id"] in agents else CONFIG.default_agent
    state.active_model = conv_row["model"]
    state.active_project_id = conv_row["project_id"]
    state.messages = await run.io_bound(memory.get_conversation_messages, conv_row["id"])
    state.title_set = True
    state.open_docs = []
    state.active_doc_path = None
    await _apply_active_project_scope(state)


# --- persistence helpers -----------------------------------------------------

def _lookup_tool_name(messages: list[dict], tool_call_id: str) -> str | None:
    for msg in messages:
        if msg.get("role") == "assistant" and msg.get("tool_calls"):
            for call in msg["tool_calls"]:
                if call.get("id") == tool_call_id:
                    return call["function"]["name"]
    return None


def persist_new_messages(conv_id: int, messages: list[dict], start_index: int) -> None:
    for msg in messages[start_index:]:
        role = msg["role"]
        tool_call_id = msg.get("tool_call_id")
        tool_name = _lookup_tool_name(messages, tool_call_id) if role == "tool" and tool_call_id else None
        memory.add_message(
            conv_id, role, msg.get("content"),
            tool_calls=msg.get("tool_calls"), tool_call_id=tool_call_id, tool_name=tool_name,
        )


def build_system_message(agent: Agent, user_input: str, skills: list[Skill]) -> dict:
    matched: list[Skill] = match_skills(user_input, skills)
    text = agent.system_prompt + "\n\n" + _AGENT_GUARDRAILS

    roots = get_active_roots()
    if roots:
        roots_str = ", ".join(str(r) for r in roots)
        text += (
            f"\n\nThe folder you can currently read/write is: {roots_str}. "
            "When the user refers to \"this folder\"/\"esta carpeta\" or gives no path, "
            "use exactly this folder."
        )

    text += "\n\n" + _time_context()

    location_ctx = _location_context()
    if location_ctx:
        text += "\n\n" + location_ctx

    if matched:
        text += "\n\n# Relevant skill instructions:\n" + "\n\n".join(
            f"## {s.name}\n{s.body}" for s in matched
        )
    return {"role": "system", "content": text}


# --- canvas (document preview panel) -----------------------------------------

def _parse_excel(raw: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header_row)]
    rows = []
    for row in rows_iter:
        rows.append({header[i]: row[i] if i < len(row) else None for i in range(len(header))})
    return rows


def _load_into_canvas(state: AppState, path: str) -> None:
    """Loads (or refreshes, if already open) `path` into state.open_docs and makes it
    the active tab - never replaces other open docs, unlike the single-slot canvas this
    replaced."""
    ext = Path(path).suffix.lower()
    try:
        if ext in _IMAGE_EXTS or ext in _PDF_EXTS:
            raw = read_file_bytes(path)
            mime = _MIME_BY_EXT.get(ext, "application/octet-stream")
            data_uri = f"data:{mime};base64," + base64.b64encode(raw).decode("ascii")
            kind = "pdf" if ext in _PDF_EXTS else "image"
            doc = {"path": path, "kind": kind, "data_uri": data_uri}
        elif ext in _EXCEL_EXTS:
            raw = read_file_bytes(path)
            doc = {"path": path, "kind": "excel", "rows": _parse_excel(raw)}
        else:
            fresh = read_file(path)
            doc = {"path": path, "kind": "text", "content": fresh["content"]}
    except Exception as e:
        doc = {"path": path, "kind": "error", "error": str(e)}

    existing_idx = next((i for i, d in enumerate(state.open_docs) if d["path"] == path), None)
    if existing_idx is not None:
        state.open_docs[existing_idx] = doc
    else:
        state.open_docs.append(doc)
    state.active_doc_path = path


def _scan_for_canvas_update(state: AppState, messages: list[dict], start_index: int) -> bool:
    """Returns True if a doc was (re)loaded, so the caller knows to reveal the panel."""
    loaded = False
    for msg in messages[start_index:]:
        if msg.get("role") != "tool":
            continue
        tool_name = _lookup_tool_name(messages, msg.get("tool_call_id"))
        if tool_name not in ("read_file", "write_file"):
            continue
        try:
            result = json.loads(msg["content"])
        except (json.JSONDecodeError, TypeError):
            continue
        path = result.get("path")
        if path and "error" not in result:
            _load_into_canvas(state, path)
            loaded = True
    return loaded


def _reveal_canvas(refs: UIRefs) -> None:
    """Forces the panel open - called only from the explicit "open a doc" actions
    (chat tool result, tree click, recent-doc click), never from a routine
    re-render, so it doesn't fight a manual collapse while docs stay open."""
    refs.content_splitter.value = _CANVAS_EXPANDED


def _switch_doc(state: AppState, refs: UIRefs, path: str) -> None:
    state.active_doc_path = path
    render_canvas(state, refs)


def _close_doc(state: AppState, refs: UIRefs, path: str) -> None:
    state.open_docs = [d for d in state.open_docs if d["path"] != path]
    if state.active_doc_path == path:
        state.active_doc_path = state.open_docs[-1]["path"] if state.open_docs else None
    render_canvas(state, refs)


def render_canvas(state: AppState, refs: UIRefs) -> None:
    container = refs.canvas_container
    container.clear()

    # Empty is always collapsed - but non-empty is left alone here (not forced
    # open), so a manual collapse while docs stay open survives routine
    # re-renders. Opening/switching to a doc explicitly calls _reveal_canvas.
    if not state.open_docs:
        refs.content_splitter.value = _CANVAS_COLLAPSED

    with container:
        with ui.row().classes("items-center justify-between w-full"):
            ui.label("Documentos").classes("text-lg font-bold")
            ui.button(icon="keyboard_double_arrow_right", on_click=lambda: _toggle_canvas(refs)) \
                .props("flat dense round").tooltip("Mostrar/ocultar Documentos")

        docs = memory.list_project_documents(state.active_project_id)
        if docs:
            with ui.expansion("Recientes en este proyecto", value=not state.open_docs).classes("w-full"):
                for path in docs:
                    ui.button(
                        Path(path).name,
                        on_click=lambda p=path: asyncio.create_task(_open_recent_doc(state, refs, p)),
                    ).props("flat align=left").classes("w-full justify-start")

        if not state.open_docs:
            ui.label("Acá vas a ver los archivos que Lechu lea o escriba durante la conversación.") \
                .classes("text-caption text-gray-500")
            return

        with ui.row().classes("lechu-doc-tabs w-full no-wrap gap-0"):
            for doc in state.open_docs:
                path = doc["path"]
                is_active = path == state.active_doc_path
                classes = "lechu-doc-tab" + (" lechu-doc-tab--active" if is_active else "")
                with ui.row().classes(classes).style("width: auto"):
                    ui.label(Path(path).name).classes("cursor-pointer").tooltip(path) \
                        .on("click", lambda p=path: _switch_doc(state, refs, p))
                    ui.icon("close", size="14px").classes("lechu-doc-tab-close cursor-pointer") \
                        .on("click", lambda p=path: _close_doc(state, refs, p))

        active = next((d for d in state.open_docs if d["path"] == state.active_doc_path), None)
        if active is None:
            return

        ui.label(active["path"]).classes("text-caption text-gray-500")
        kind = active.get("kind", "text")
        if kind == "error":
            ui.label(active["error"]).classes("text-red")
            return

        if kind == "image":
            ui.image(active["data_uri"]).classes("w-full")
        elif kind == "pdf":
            # sanitize=False: ui.html defaults to client-side DOMPurify sanitization,
            # which silently strips <embed> (not in its allowed-tags list) - the
            # data URI is our own base64 of a file we just read, not user input.
            # #view=FitH: Chromium's built-in PDF viewer otherwise opens at a fixed
            # zoom level instead of filling the available width.
            ui.html(
                f'<embed src="{active["data_uri"]}#view=FitH" type="application/pdf" '
                'style="width:100%; height:calc(100vh - 220px); border:none;" />',
                sanitize=False,
            ).classes("w-full")
        elif kind == "excel":
            rows = active["rows"]
            if rows:
                columns = [{"name": k, "label": k, "field": k} for k in rows[0].keys()]
                ui.table(rows=rows, columns=columns, row_key=list(rows[0].keys())[0]).classes("w-full")
            else:
                ui.label("(Excel vacío)")
        else:
            content = active["content"]
            ext = Path(active["path"]).suffix.lower()
            if ext in (".csv", ".tsv"):
                delimiter = "," if ext == ".csv" else "\t"
                try:
                    rows = list(csv.DictReader(io.StringIO(content), delimiter=delimiter))
                    if rows:
                        columns = [{"name": k, "label": k, "field": k} for k in rows[0].keys()]
                        ui.table(rows=rows, columns=columns, row_key=list(rows[0].keys())[0]).classes("w-full")
                    else:
                        ui.label("(CSV vacío)")
                except csv.Error:
                    ui.code(content)
            elif ext in (".md", ".markdown"):
                ui.markdown(content)
            else:
                ui.code(content, language=_guess_language(ext)).classes("w-full")


async def _open_recent_doc(state: AppState, refs: UIRefs, path: str) -> None:
    await run.io_bound(_load_into_canvas, state, path)
    render_canvas(state, refs)
    _reveal_canvas(refs)


# --- explorer (file tree) -----------------------------------------------------

_DESIGN_EXTS = {".fig", ".sketch", ".psd", ".ai", ".xd"}
_CFG_EXTS = {".yaml", ".yml", ".toml", ".ini", ".cfg", ".env", ".conf"}

_TREE_HEADER_SLOT = """
    <div class="row items-center no-wrap full-width lechu-tree-row">
      <img :src="props.node.icon" width="16" height="16" class="q-mr-xs" />
      <div class="ellipsis">{{ props.node.label }}</div>
      <q-badge v-if="props.node.count" class="lechu-tree-badge q-ml-sm" :label="props.node.count" />
    </div>
"""


def _icon_for_entry(entry: Path) -> str:
    if entry.is_dir():
        return _tabler_icon("folder")
    ext = entry.suffix.lower()
    if ext == ".js":
        return _tabler_icon("js")
    if ext == ".jsx":
        return _tabler_icon("jsx")
    if ext == ".css":
        return _tabler_icon("css")
    if ext == ".json":
        return _tabler_icon("json")
    if ext in (".md", ".markdown"):
        return _tabler_icon("md")
    if ext in _IMAGE_EXTS:
        return _tabler_icon("image")
    if ext in _DESIGN_EXTS:
        return _tabler_icon("design")
    if ext in (".csv", ".tsv") or ext in _EXCEL_EXTS:
        return _tabler_icon("sheet")
    if ext in _CFG_EXTS:
        return _tabler_icon("cfg")
    if ext in _PDF_EXTS:
        return _tabler_icon("pdf")
    if ext in _LANG_BY_EXT:
        return _tabler_icon("code")
    return _tabler_icon("default")


def _build_tree_nodes(folder: Path, depth: int = 0, max_entries: int = 300, max_depth: int = 8) -> list[dict]:
    if depth > max_depth:
        return []
    try:
        entries = sorted(folder.iterdir(), key=lambda p: (p.is_file(), p.name.lower()))
    except OSError:
        return []
    entries = [e for e in entries if not e.name.startswith(".")][:max_entries]
    nodes = []
    for entry in entries:
        node: dict = {"id": str(entry), "label": entry.name, "icon": _icon_for_entry(entry)}
        if entry.is_dir():
            children = _build_tree_nodes(entry, depth + 1, max_entries, max_depth)
            node["children"] = children
            node["count"] = len(children)
        nodes.append(node)
    return nodes


def _collect_expand_ids(nodes: list[dict], query: str, ancestors: tuple[str, ...] = ()) -> set[str]:
    """Ids of every folder that needs to be force-expanded to reveal a search match
    somewhere below it - ui.tree's own `filter` already hides/shows rows, but doesn't
    auto-expand collapsed ancestors of a match."""
    result: set[str] = set()
    for node in nodes:
        children = node.get("children")
        self_matches = query in node["label"].lower()
        child_ids = _collect_expand_ids(children, query, ancestors + (node["id"],)) if children else set()
        if self_matches or child_ids:
            result.update(ancestors)
        result |= child_ids
    return result


def _active_base_folder(state: AppState) -> Path | None:
    project = _active_project_row(state)
    if project and project["folder_path"]:
        return Path(project["folder_path"])
    roots = get_active_roots()
    return roots[0] if roots else None


async def _prompt_dialog(message: str) -> str | None:
    with ui.dialog() as dialog, ui.card():
        ui.label(message)
        name_input = ui.input().classes("w-full")
        name_input.on("keydown.enter", lambda: dialog.submit(name_input.value))
        with ui.row().classes("w-full justify-end"):
            ui.button("Cancelar", on_click=lambda: dialog.submit(None))
            ui.button("Crear", on_click=lambda: dialog.submit(name_input.value), color="primary")
    return await dialog


async def _new_folder_dialog(state: AppState, refs: UIRefs) -> None:
    name = await _prompt_dialog("Nombre de la nueva carpeta:")
    if not name:
        return
    base = _active_base_folder(state)
    if base is None:
        ui.notify("Abrí una carpeta primero.", type="warning")
        return
    try:
        await run.io_bound(create_folder, str(base / name))
        render_explorer(state, refs)
    except Exception as e:
        ui.notify(str(e), type="negative")


async def _new_file_dialog(state: AppState, refs: UIRefs) -> None:
    name = await _prompt_dialog("Nombre del nuevo archivo:")
    if not name:
        return
    base = _active_base_folder(state)
    if base is None:
        ui.notify("Abrí una carpeta primero.", type="warning")
        return
    try:
        await run.io_bound(write_file, str(base / name), "")
        render_explorer(state, refs)
    except Exception as e:
        ui.notify(str(e), type="negative")


def render_explorer(state: AppState, refs: UIRefs) -> None:
    container = refs.files_container
    container.clear()
    tree_holder: dict = {}
    with container:
        with ui.row().classes("items-center justify-between w-full q-px-sm q-pt-sm"):
            ui.label("EXPLORADOR").classes("lechu-section-label")
            with ui.row().classes("items-center gap-0"):
                ui.button(icon="create_new_folder",
                          on_click=lambda: asyncio.create_task(_new_folder_dialog(state, refs))) \
                    .props("flat dense round").tooltip("Nueva carpeta")
                ui.button(icon="note_add",
                          on_click=lambda: asyncio.create_task(_new_file_dialog(state, refs))) \
                    .props("flat dense round").tooltip("Nuevo archivo")
                ui.button(icon="unfold_less",
                          on_click=lambda: tree_holder["tree"].collapse()) \
                    .props("flat dense round").tooltip("Colapsar todo")

        project = _active_project_row(state)
        if not (project and project["folder_path"]):
            ui.label("Abrí una carpeta desde arriba para ver sus archivos acá.") \
                .classes("text-caption text-gray-500 q-px-sm")
            return

        folder = Path(project["folder_path"])
        nodes = _build_tree_nodes(folder)

        search_input = ui.input(placeholder="Buscar archivos...") \
            .props("dense borderless").classes("w-full lechu-search q-px-sm")
        with search_input.add_slot("prepend"):
            ui.icon("search").classes("text-sm")

        if not nodes:
            ui.label("(carpeta vacía)").classes("text-caption text-gray-500 q-px-sm")
            return

        tree = ui.tree(nodes, node_key="id", label_key="label",
                        on_select=lambda e: _on_tree_select(state, refs, e.value)) \
            .props("no-connectors dense").classes("lechu-tree w-full")
        tree.add_slot("default-header", _TREE_HEADER_SLOT)
        tree_holder["tree"] = tree

        def _on_files_search_change(e) -> None:
            text = (e.value or "").strip()
            tree.set_filter(text)
            if text:
                tree.expand(list(_collect_expand_ids(nodes, text.lower())))

        search_input.on_value_change(_on_files_search_change)


def _on_tree_select(state: AppState, refs: UIRefs, node_id: str | None) -> None:
    if not node_id:
        return
    if Path(node_id).is_file():
        asyncio.create_task(_open_recent_doc(state, refs, node_id))


# --- chat rendering -----------------------------------------------------------

def render_chat_history(state: AppState, container: ui.column) -> None:
    """Full replay of a conversation. Tool-call/tool-result messages are internal
    plumbing, not shown - only real user/assistant text bubbles are rendered."""
    container.clear()
    with container:
        for msg in state.messages:
            role = msg["role"]
            if role == "user":
                with ui.chat_message(sent=True):
                    ui.markdown(msg["content"])
            elif role == "assistant" and not msg.get("tool_calls"):
                with ui.chat_message(name="Lechu", avatar=OWL_AVATAR):
                    ui.markdown(msg["content"])


def _new_thinking_bubble(container: ui.column, agent: Agent) -> tuple[ui.element, ui.row, ui.markdown]:
    """Creates one chat bubble for the whole turn, showing an animated 'Pensando...'
    placeholder that gets swapped for real content once it starts streaming in."""
    with container:
        with ui.chat_message(name=agent.name, avatar=OWL_AVATAR) as bubble:
            thinking = ui.row().classes("items-center gap-2")
            with thinking:
                ui.html('<span class="lechu-thinking-owl">🦉</span>')
                ui.label("Pensando...").classes("italic text-gray-500")
            md = ui.markdown("")
            md.set_visibility(False)
    return bubble, thinking, md


async def _stream_into_chat(
    md: ui.markdown, thinking: ui.row, chunks: AsyncIterator[str],
    min_interval: float = 0.08, min_chars: int = 24,
) -> bool:
    """Fills `md` with streamed content, swapping out the `thinking` placeholder
    the moment real content starts arriving. Returns True if any content streamed
    (tool-only turns yield no content at all)."""
    buf: list[str] = []
    total_len = flushed_len = 0
    loop = asyncio.get_event_loop()
    last_flush = loop.time()
    started = False

    async for piece in chunks:
        if not started:
            started = True
            thinking.set_visibility(False)
            md.set_visibility(True)
        buf.append(piece)
        total_len += len(piece)
        now = loop.time()
        if total_len - flushed_len >= min_chars or now - last_flush >= min_interval:
            md.set_content("".join(buf))
            flushed_len, last_flush = total_len, now

    if started:
        md.set_content("".join(buf))
    return started


# --- turn driving --------------------------------------------------------

def _summarize_tool_call(tool_name: str, args: dict) -> str:
    if tool_name == "write_file":
        return f"Quiere escribir el archivo {args.get('path', '?')}"
    if tool_name == "delete_file":
        return f"Quiere borrar el archivo {args.get('path', '?')}"
    if tool_name == "write_drive_file":
        action = "sobrescribir" if args.get("file_id") else "crear"
        return f"Quiere {action} el archivo de Drive \"{args.get('name', '?')}\""
    if tool_name == "send_email":
        return f"Quiere enviar un mail a {args.get('to', '?')} con asunto \"{args.get('subject', '?')}\""
    if tool_name == "manage_calendar_event":
        action = args.get("action")
        title = args.get("title", "?")
        if action == "create":
            color = f" (color {args['color']})" if args.get("color") else ""
            guests = f" e invitar a {', '.join(args['attendees'])}" if args.get("attendees") else ""
            return f"Quiere crear el evento \"{title}\" el {args.get('start', '?')}{color}{guests}"
        if action == "update":
            return f"Quiere modificar el evento {args.get('event_id', '?')}"
        if action == "delete":
            return f"Quiere borrar el evento {args.get('event_id', '?')}"
    return f"Quiere ejecutar {tool_name}"


async def _ask_confirmation(container: ui.column, agent: Agent, pc: PendingConfirmation) -> dict:
    future: asyncio.Future = asyncio.get_event_loop().create_future()
    holder: dict = {}

    def _resolve(approved: bool) -> None:
        if future.done():
            return
        row = holder["row"]
        row.clear()
        with row:
            ui.label("✅ Confirmado" if approved else "❌ Cancelado").classes("text-xs text-gray-500")
        future.set_result(approved)

    with container:
        with ui.chat_message(name=agent.name, avatar=OWL_AVATAR):
            ui.label(_summarize_tool_call(pc.tool_name, pc.args)).classes("text-orange-700 font-medium")
            with ui.row() as btn_row:
                ui.button("Confirmar", on_click=lambda: _resolve(True), color="primary")
                ui.button("Cancelar", on_click=lambda: _resolve(False))
            holder["row"] = btn_row

    approved = await future

    if approved:
        result = await run.io_bound(execute_tool, pc.tool_name, pc.args)
    else:
        result = {"error": "El usuario rechazó esta acción."}
    return result if result is not None else {"error": "cancelled"}


async def _advance(
    state: AppState, refs: UIRefs, agent: Agent, failed_counts: dict[tuple, int] | None = None,
) -> None:
    # Owned for the whole user turn, not just this call - passed through the
    # recursive call below (after a confirmation resolves) so a repeated
    # failing call is still caught across that boundary.
    if failed_counts is None:
        failed_counts = {}
    start_len = len(state.messages)
    bubble, thinking, md = _new_thinking_bubble(refs.chat_container, agent)
    try:
        result = None
        content_started = False
        for _ in range(CONFIG.max_tool_iterations):
            chunks, box = await step_agent_stream(get_client(), agent, state.messages, failed_counts)
            if await _stream_into_chat(md, thinking, chunks):
                content_started = True
            result = box["result"]
            if isinstance(result, Continue):
                # Reveal a newly-read/written doc as soon as the tool result lands,
                # not only once the whole turn (including the model's follow-up
                # commentary) finishes - for a big file that commentary can take a
                # while, and the tab shouldn't make the user wait for it too.
                if _scan_for_canvas_update(state, state.messages, start_len):
                    render_canvas(state, refs)
                    _reveal_canvas(refs)
                continue
            break
        else:
            fallback = "Me detuve después de demasiadas llamadas a herramientas. Probá simplificar el pedido."
            state.messages.append({"role": "assistant", "content": fallback})
            thinking.set_visibility(False)
            md.set_visibility(True)
            md.set_content(fallback)
            content_started = True
            result = FinalAnswer(fallback)

        if isinstance(result, PendingConfirmation) and not content_started:
            bubble.delete()

        await run.io_bound(persist_new_messages, state.conversation_id, state.messages, start_len)
        canvas_loaded = _scan_for_canvas_update(state, state.messages, start_len)
        render_canvas(state, refs)
        if canvas_loaded:
            _reveal_canvas(refs)

        if isinstance(result, PendingConfirmation):
            tool_result = await _ask_confirmation(refs.chat_container, agent, result)
            pre_append_len = len(state.messages)
            state.messages.append({
                "role": "tool", "tool_call_id": result.tool_call_id, "content": json.dumps(tool_result),
            })
            await run.io_bound(persist_new_messages, state.conversation_id, state.messages, pre_append_len)
            if result.tool_name in ("read_file", "write_file") and "error" not in tool_result:
                await run.io_bound(_load_into_canvas, state, tool_result["path"])
                render_canvas(state, refs)
                _reveal_canvas(refs)
                render_explorer(state, refs)
            await _advance(state, refs, agent, failed_counts)

    except httpx.HTTPError as e:
        msg = f"No se pudo contactar a Ollama: {e}"
        state.messages.append({"role": "assistant", "content": msg})
        thinking.set_visibility(False)
        md.set_visibility(True)
        md.set_content(msg)
        await run.io_bound(persist_new_messages, state.conversation_id, state.messages, start_len)


async def handle_user_turn(state: AppState, refs: UIRefs, agent: Agent, user_input: str) -> None:
    state.messages[0] = build_system_message(agent, user_input, state.skills)
    state.messages.append({"role": "user", "content": user_input})
    await run.io_bound(memory.add_message, state.conversation_id, "user", user_input)

    if not state.title_set:
        title = user_input.strip().splitlines()[0][:60]
        await run.io_bound(memory.set_conversation_title, state.conversation_id, title)
        state.title_set = True

    with refs.chat_container:
        with ui.chat_message(sent=True):
            ui.markdown(user_input)

    await _advance(state, refs, agent)


async def _on_submit(state: AppState, refs: UIRefs, agents: dict[str, Agent], refs_holder: dict) -> None:
    text = (refs.input_box.value or "").strip()
    if not text or (state.current_turn_task and not state.current_turn_task.done()):
        return
    refs.input_box.value = ""
    refs.input_box.disable()
    refs.send_btn.disable()

    routed_agent_id = await run.io_bound(route_agent, text, list(agents.keys()), state.agent_id)
    if routed_agent_id != state.agent_id:
        state.agent_id = routed_agent_id
        state.active_model = agents[routed_agent_id].model
        refs_holder["agent_select"].value = routed_agent_id
        refs_holder["model_select"].value = state.active_model

    agent = _effective_agent(state)
    task = asyncio.create_task(handle_user_turn(state, refs, agent, text))
    state.current_turn_task = task
    try:
        await task
    except asyncio.CancelledError:
        pass
    finally:
        state.current_turn_task = None
        refs.input_box.enable()
        refs.send_btn.enable()


# --- date bucketing (Historial grouping) ---------------------------------------

_WEEKDAY_ES_ABBR = {0: "Lun", 1: "Mar", 2: "Mié", 3: "Jue", 4: "Vie", 5: "Sáb", 6: "Dom"}
_BUCKET_ORDER = ("Hoy", "Ayer", "Esta semana", "Más antiguo")


def _to_local_dt(updated_at: str) -> datetime | None:
    """`updated_at` is stored as SQLite's `datetime('now')`, i.e. UTC - convert to
    local time before bucketing, otherwise "Hoy"/"Ayer" drift near midnight."""
    try:
        dt_utc = datetime.strptime(updated_at, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
    except (ValueError, TypeError):
        return None
    return dt_utc.astimezone()


def _bucket_for(dt_local: datetime | None, today_local: date) -> str:
    if dt_local is None:
        return "Más antiguo"
    d = dt_local.date()
    if d == today_local:
        return "Hoy"
    if d == today_local - timedelta(days=1):
        return "Ayer"
    if d >= today_local - timedelta(days=6):
        return "Esta semana"
    return "Más antiguo"


def _format_timestamp(dt_local: datetime | None, bucket: str) -> str:
    if dt_local is None:
        return ""
    if bucket == "Hoy":
        return dt_local.strftime("%H:%M")
    if bucket in ("Ayer", "Esta semana"):
        return _WEEKDAY_ES_ABBR[dt_local.weekday()]
    return dt_local.strftime("%d/%m")


# --- main page -----------------------------------------------------------------

def build_page() -> None:
    @ui.page("/")
    async def main_page() -> None:
        ui.colors(primary="#a5693a", secondary="#e8a33d", accent="#c98f52")
        ui.add_head_html("""
            <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Nunito:wght@500;700;800&display=swap">
            <style>
            @keyframes lechu-wiggle {
                0%, 100% { transform: rotate(-12deg); }
                50% { transform: rotate(12deg); }
            }
            .lechu-thinking-owl {
                display: inline-block;
                font-size: 1.3em;
                transform-origin: 50% 80%;
            }
            body.lechu-wiggle-on .lechu-thinking-owl {
                animation: lechu-wiggle 0.8s ease-in-out infinite;
            }

            /* brandbook theme tokens - light (default) and dark ("dos temas, una madera").
               "Acogedor y rico": deeper wood tones, warm gradients, generous
               rounding - values sourced from the "Lechu - Direcciones de
               diseño" mockup, Option C, chosen by the user over A/B. */
            :root {
                --surface-0: #f7ede0;
                --surface-1: #fbf4e8;
                --surface-2: #ead9bc;
                --text-accent: #a5693a;
                --text-accent-2: #8f5a30;
                --bg-accent: #e8a33d;
                --bg-accent-text: #ffffff;
                --text-primary: #3d2b1a;
                --text-secondary: rgba(61, 43, 26, 0.75);
                --text-muted: rgba(61, 43, 26, 0.62);
                --border: rgba(90, 58, 30, 0.14);
                --radius-sm: 12px;
                --radius-md: 16px;
                --radius-lg: 20px;
                --sidebar-gradient: linear-gradient(180deg, #f3e4cd 0%, #ecd7b6 100%);
                --card-shadow: 0 1px 3px rgba(90, 58, 30, 0.12);
            }
            body.body--dark {
                --surface-0: #2a1a10;
                --surface-1: #3a2717;
                --surface-2: #4a3220;
                --text-accent: #e8a33d;
                --text-accent-2: #f0b85e;
                --bg-accent: #e8a33d;
                --bg-accent-text: #2a1a10;
                --text-primary: #f5e9d8;
                --text-secondary: rgba(245, 233, 216, 0.75);
                --text-muted: rgba(245, 233, 216, 0.6);
                --border: rgba(245, 233, 216, 0.14);
                --sidebar-gradient: linear-gradient(180deg, #3a2717 0%, #2a1a10 100%);
                --card-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
            }
            body {
                background: var(--surface-0);
                font-size: 12.5px;
                font-family: 'Nunito', system-ui, sans-serif;
            }
            .q-drawer { background: var(--sidebar-gradient); }

            /* native Quasar inputs/buttons pick up the warmer radius without
               touching every individual ui.select/ui.input/ui.button call site */
            .q-field--outlined .q-field__control { border-radius: var(--radius-md) !important; }
            .q-btn { border-radius: var(--radius-sm); }

            /* Quasar's sent bubble paints via the theme's primary color, which
               can't carry a gradient - overridden explicitly here. The received
               bubble still uses the `background: currentColor` trick (setting
               `color` on the wrapper repaints the bubble), just with the new
               warm fill instead of Quasar's default light-green. */
            .q-message-text--sent {
                background: linear-gradient(135deg, #a5693a, #c98f52) !important;
                border-radius: var(--radius-lg) var(--radius-lg) 4px var(--radius-lg) !important;
                box-shadow: 0 3px 8px rgba(165, 105, 58, 0.3);
            }
            .q-message-text-content--sent { color: #ffffff; }
            .q-message-text--received {
                color: #ead9bc;
                border-radius: var(--radius-lg) var(--radius-lg) var(--radius-lg) 4px !important;
            }
            .q-message-text-content--received { color: #3d2b1a; }
            body.body--dark .q-message-text--received { color: #4a3220; }
            body.body--dark .q-message-text-content--received { color: #f5e9d8; }

            /* font size */
            body.lechu-font-small .q-message-text-content { font-size: 0.85em; }
            body.lechu-font-medium .q-message-text-content { font-size: 1em; }
            body.lechu-font-large .q-message-text-content { font-size: 1.2em; }

            /* sidebar: list items (Historial), section labels, search boxes */
            .lechu-list-item {
                display: flex; align-items: center; gap: 8px;
                padding: 6px 8px; border-radius: var(--radius-sm);
                color: var(--text-primary); cursor: pointer;
            }
            .lechu-list-item:hover { background: var(--surface-2); }
            .lechu-list-item--active {
                background: var(--bg-accent); color: var(--bg-accent-text);
                box-shadow: 0 2px 6px rgba(232, 163, 61, 0.35);
            }
            .lechu-timestamp { font-size: 10px; color: var(--text-muted); }
            .lechu-list-item--active .lechu-timestamp { color: var(--bg-accent-text); opacity: 0.75; }

            /* canvas panel: open-document tabs (editor-style strip, not pills) */
            .lechu-doc-tabs {
                display: flex; align-items: stretch; border-bottom: 1px solid var(--border);
                overflow-x: auto; overflow-y: hidden;
            }
            .lechu-doc-tab {
                display: flex; align-items: center; gap: 6px; flex-shrink: 0;
                padding: 6px 8px 6px 10px; font-size: 12px; white-space: nowrap;
                color: var(--text-muted); background: var(--surface-1);
                border-right: 1px solid var(--border);
                border-bottom: 2px solid transparent; cursor: pointer;
            }
            .lechu-doc-tab:hover { color: var(--text-primary); background: var(--surface-2); }
            .lechu-doc-tab--active {
                color: var(--text-primary); background: var(--surface-0);
                border-bottom: 2px solid var(--text-accent); font-weight: 600;
            }
            .lechu-doc-tab-close { opacity: 0.5; border-radius: 4px; }
            .lechu-doc-tab-close:hover { opacity: 1; background: var(--surface-2); }
            .lechu-section-label {
                font-size: 10px; text-transform: uppercase; letter-spacing: .06em;
                color: var(--text-muted); font-weight: 600;
            }
            .lechu-search { background: var(--surface-1); border-radius: var(--radius-md); box-shadow: var(--card-shadow); }

            /* sidebar: file tree */
            .lechu-tree .q-tree__node { padding-left: 14px; }
            .lechu-tree .q-tree__children { padding-left: 14px; }
            .lechu-tree .q-tree__node-header { border-radius: var(--radius-sm); }
            .lechu-tree .q-tree__node-header:hover { background: var(--surface-2); }
            /* Quasar hardcodes a grey (#9e9e9e) for selected-node text - override
               or our accent color silently loses to it. */
            .lechu-tree .q-tree__node--selected .q-tree__node-header { background: var(--bg-accent); }
            .lechu-tree .q-tree__node--selected .q-tree__node-header-content { color: var(--bg-accent-text) !important; }
            .lechu-tree-badge { background: var(--surface-2); color: var(--text-muted); font-size: 10px; }
            </style>
        """)

        initial_font_size = memory.get_setting("font_size", "medium")
        initial_wiggle = memory.get_setting("wiggle", "on")
        ui.query("body").classes(
            f"lechu-font-{initial_font_size} " + ("lechu-wiggle-on" if initial_wiggle == "on" else "")
        )

        agents = get_agents()
        if not agents:
            ui.label(f"No hay agentes definidos en {AGENTS_DIR}. Agregá al menos un archivo .yaml.").classes("text-red")
            return

        default_agent_id = CONFIG.default_agent if CONFIG.default_agent in agents else next(iter(agents))
        state = AppState(agent_id=default_agent_id, active_model=agents[default_agent_id].model)
        state.skills = await run.io_bound(load_skills, SKILLS_DIR)
        await start_new_conversation(state, agents[state.agent_id])
        await _apply_active_project_scope(state)

        try:
            available_models = await get_client().list_models()
        except httpx.HTTPError:
            available_models = []
        if available_models and state.active_model not in available_models:
            state.active_model = available_models[0]

        refs_holder: dict = {}

        with ui.header().classes("items-center justify-between bg-[#a5693a]"):
            ui.label("🦉 Lechu").classes("text-xl font-bold")

            @ui.refreshable
            def header_status_label() -> None:
                weather = _header_weather.get("data")
                text = (
                    f"{_format_header_date()} · {weather['temperature_c']:.0f}°C {weather['emoji']}"
                    if weather else _format_header_date()
                )
                ui.label(text).classes("text-sm opacity-80")

            header_status_label()

            async def _refresh_header_weather() -> None:
                loc = location_service.get_cached_location()
                if loc:
                    result = await run.io_bound(get_weather, loc["city"])
                    if "error" not in result:
                        _header_weather["data"] = result["current"]
                header_status_label.refresh()

            asyncio.create_task(_refresh_header_weather())
            ui.timer(3600, lambda: asyncio.create_task(_refresh_header_weather()))

        # width=250 via .props(), not .style(): Quasar's QDrawer uses its own
        # `width` prop both for the CSS width AND for the page-content margin
        # it reserves - a pure .style() override desyncs those two and causes
        # overlap/gap between the drawer and the content splitter next to it.
        with ui.left_drawer(value=True, bordered=True).props("width=250").classes("column no-wrap") as drawer:
            files_container = render_sidebar(state, agents, available_models, refs_holder)

        with ui.splitter(value=_CANVAS_COLLAPSED).classes("w-full") \
                .style("height: calc(100vh - 64px)") as content_splitter:
            with content_splitter.before:
                with ui.column().classes("w-full h-full p-2"):
                    chat_container = ui.column().classes("w-full flex-grow overflow-auto")

                    async def _on_agent_change(agent_id: str) -> None:
                        state.agent_id = agent_id
                        state.active_model = agents[agent_id].model
                        await start_new_conversation(state, _effective_agent(state))
                        refs = refs_holder["refs"]
                        render_chat_history(state, refs.chat_container)
                        render_canvas(state, refs)
                        refs_holder["refresh_history"]()
                        refs_holder["model_select"].value = state.active_model

                    def _on_model_change(model: str) -> None:
                        state.active_model = model

                    with ui.row().classes("w-full items-center gap-2"):
                        agent_select = ui.select(
                            {aid: a.name for aid, a in agents.items()}, value=state.agent_id,
                            on_change=lambda e: asyncio.create_task(_on_agent_change(e.value)),
                        ).props("dense outlined options-dense").classes("w-36")
                        refs_holder["agent_select"] = agent_select

                        model_options = available_models if available_models else [state.active_model]
                        model_select = ui.select(
                            model_options, value=state.active_model,
                            on_change=lambda e: _on_model_change(e.value),
                        ).props("dense outlined options-dense").classes("w-28")
                        refs_holder["model_select"] = model_select

                        input_box = ui.input(placeholder="Escribí un mensaje...").classes("flex-grow")
                        send_btn = ui.button(icon="send")
            with content_splitter.after:
                canvas_container = ui.column().classes("w-full h-full overflow-auto p-2")

        refs = UIRefs(
            chat_container=chat_container,
            files_container=files_container,
            canvas_container=canvas_container,
            input_box=input_box,
            send_btn=send_btn,
            content_splitter=content_splitter,
        )
        refs_holder["refs"] = refs

        input_box.on("keydown.enter", lambda: asyncio.create_task(_on_submit(state, refs, agents, refs_holder)))
        send_btn.on_click(lambda: asyncio.create_task(_on_submit(state, refs, agents, refs_holder)))

        render_chat_history(state, chat_container)
        render_explorer(state, refs)
        render_canvas(state, refs)


async def _confirm_dialog(message: str) -> bool:
    with ui.dialog() as dialog, ui.card():
        ui.label(message)
        with ui.row().classes("w-full justify-end"):
            ui.button("Cancelar", on_click=lambda: dialog.submit(False))
            ui.button("Confirmar", on_click=lambda: dialog.submit(True), color="red")
    return bool(await dialog)


# --- sidebar -------------------------------------------------------------------

def render_sidebar(state: AppState, agents: dict[str, Agent], available_models: list[str], refs_holder: dict) -> ui.column:
    # --- pinned block: Proyecto/Agente/Modelo affect both tabs, always visible ---
    with ui.column().classes("w-full gap-1 q-pa-sm"):
        @ui.refreshable
        def project_selector() -> None:
            project_rows = memory.list_projects()
            options = {None: "Sin proyecto"}
            for p in project_rows:
                options[p["id"]] = p["name"]
            with ui.row().classes("w-full items-center gap-1 no-wrap"):
                ui.select(options, value=state.active_project_id, label="Proyecto",
                          on_change=lambda e: asyncio.create_task(_on_project_change(e.value))) \
                    .props("dense outlined options-dense").classes("flex-grow")
                ui.button(icon="folder_open", on_click=lambda: asyncio.create_task(_open_folder())) \
                    .props("flat dense round").tooltip("Abrir carpeta...")

        project_selector()
        refs_holder["refresh_project_selector"] = project_selector.refresh

        async def _on_project_change(project_id: int | None) -> None:
            state.active_project_id = project_id
            await _apply_active_project_scope(state)
            await start_new_conversation(state, _effective_agent(state))
            refs = refs_holder["refs"]
            render_chat_history(state, refs.chat_container)
            render_explorer(state, refs)
            render_canvas(state, refs)
            refs_holder["refresh_history"]()
            folder_caption.refresh()

        async def _open_folder() -> None:
            result = await app.native.main_window.create_file_dialog(dialog_type=webview.FileDialog.FOLDER)
            if not result:
                return
            path = result[0] if isinstance(result, (list, tuple)) else result
            project_id = await run.io_bound(open_folder_as_project, path)
            state.active_project_id = project_id
            await _apply_active_project_scope(state)
            await start_new_conversation(state, _effective_agent(state))
            refs = refs_holder["refs"]
            project_selector.refresh()
            render_chat_history(state, refs.chat_container)
            render_explorer(state, refs)
            render_canvas(state, refs)
            refs_holder["refresh_history"]()
            folder_caption.refresh()

        @ui.refreshable
        def folder_caption() -> None:
            project = _active_project_row(state)
            if project and project["folder_path"]:
                ui.label(project["folder_path"]).classes("lechu-timestamp break-all")

        folder_caption()
        refs_holder["refresh_folder_caption"] = folder_caption.refresh

        if not available_models:
            ui.label(f"No se pudo conectar con Ollama en {CONFIG.ollama_base_url}.").classes("text-red text-caption")

    # --- tabs: Chats | Archivos ---
    with ui.tabs().classes("w-full") as sidebar_tabs:
        ui.tab("chats", label="Chats", icon="chat")
        ui.tab("archivos", label="Archivos", icon="folder")

    files_container_holder: dict = {}
    with ui.tab_panels(sidebar_tabs, value="chats").classes("w-full flex-grow").style("overflow: hidden"):
        with ui.tab_panel("chats").classes("h-full overflow-auto q-pa-xs"):
            render_chats_tab(state, agents, refs_holder)
        with ui.tab_panel("archivos").classes("h-full overflow-auto q-pa-none"):
            files_container_holder["container"] = ui.column().classes("w-full h-full")

    # --- footer: fixed, not scrolling ---
    with ui.row().classes("w-full items-center q-pa-sm").style("border-top: 1px solid var(--border)"):
        ui.button("Configuración", icon="settings",
                  on_click=lambda: _open_settings_dialog(state, agents, refs_holder)) \
            .props("flat align=left").classes("w-full justify-start")

    return files_container_holder["container"]


def render_chats_tab(state: AppState, agents: dict[str, Agent], refs_holder: dict) -> None:
    search_state = {"query": ""}

    async def _new_conversation() -> None:
        await start_new_conversation(state, _effective_agent(state))
        refs = refs_holder["refs"]
        render_chat_history(state, refs.chat_container)
        render_canvas(state, refs)
        history_panel.refresh()

    ui.button("Nueva conversación", icon="add", on_click=_new_conversation) \
        .props("outline dense").classes("w-full q-mb-xs text-caption")

    search_input = ui.input(placeholder="Buscar conversaciones...") \
        .props("dense borderless").classes("w-full lechu-search")
    with search_input.add_slot("prepend"):
        ui.icon("search").classes("text-sm")

    def _on_chats_search_change(e) -> None:
        search_state["query"] = (e.value or "").strip().lower()
        history_panel.refresh()

    search_input.on_value_change(_on_chats_search_change)

    @ui.refreshable
    def history_panel() -> None:
        conversations = memory.list_conversations(state.active_project_id)
        query = search_state["query"]
        if query:
            conversations = [c for c in conversations if query in (c["title"] or "").lower()]
        if not conversations:
            msg = "Sin resultados." if query else "Sin conversaciones todavía en este proyecto."
            ui.label(msg).classes("text-caption text-gray-500 q-pa-sm")
            return

        today_local = datetime.now().astimezone().date()
        buckets: dict[str, list] = {b: [] for b in _BUCKET_ORDER}
        for conv in conversations:
            dt_local = _to_local_dt(conv["updated_at"])
            buckets[_bucket_for(dt_local, today_local)].append((conv, dt_local))

        for bucket in _BUCKET_ORDER:
            items = buckets[bucket]
            if not items:
                continue
            ui.label(bucket).classes("lechu-section-label q-mt-sm q-mb-xs")
            for conv, dt_local in items:
                label = conv["title"] or f"Conversación #{conv['id']}"
                ts = _format_timestamp(dt_local, bucket)
                is_active = conv["id"] == state.conversation_id
                classes = "lechu-list-item w-full" + (" lechu-list-item--active" if is_active else "")
                with ui.row().classes(classes).on(
                    "click", lambda c=conv: asyncio.create_task(_load_history(c))
                ):
                    ui.label(label).classes("ellipsis col-grow")
                    ui.label(ts).classes("lechu-timestamp")

    async def _load_history(conv_row: sqlite3.Row) -> None:
        await load_conversation(state, conv_row, agents)
        refs = refs_holder["refs"]
        render_chat_history(state, refs.chat_container)
        render_explorer(state, refs)
        render_canvas(state, refs)
        refs_holder["refresh_project_selector"]()
        refs_holder["refresh_folder_caption"]()
        history_panel.refresh()

    history_panel()
    refs_holder["refresh_history"] = history_panel.refresh


def _open_settings_dialog(state: AppState, agents: dict[str, Agent], refs_holder: dict) -> None:
    with ui.dialog() as dialog, ui.card().classes("w-[560px] max-w-[92vw] max-h-[85vh] overflow-y-auto gap-2"):
        with ui.row().classes("w-full items-center justify-between"):
            ui.label("Configuración").classes("text-lg font-bold")
            ui.button(icon="close", on_click=dialog.close).props("flat dense round")

        with ui.expansion("Apariencia", value=True).classes("w-full"):
            theme_is_dark = memory.get_setting("theme", "light") == "dark"
            dark_mode_el = ui.dark_mode(value=theme_is_dark)

            async def _on_theme_change(value: bool) -> None:
                await run.io_bound(memory.set_setting, "theme", "dark" if value else "light")

            ui.switch(
                "Tema oscuro",
                value=theme_is_dark,
                on_change=lambda e: (dark_mode_el.set_value(e.value), asyncio.create_task(_on_theme_change(e.value))),
            )

            def _apply_font_size(size: str) -> None:
                ui.query("body").classes(
                    remove="lechu-font-small lechu-font-medium lechu-font-large",
                    add=f"lechu-font-{size}",
                )

            async def _on_font_change(e) -> None:
                _apply_font_size(e.value)
                await run.io_bound(memory.set_setting, "font_size", e.value)

            ui.select(
                {"small": "Chico", "medium": "Mediano", "large": "Grande"},
                value=memory.get_setting("font_size", "medium"),
                label="Tamaño de letra",
                on_change=_on_font_change,
            ).classes("w-full")

            async def _on_wiggle_change(e) -> None:
                ui.query("body").classes(
                    add="lechu-wiggle-on" if e.value else "", remove="" if e.value else "lechu-wiggle-on"
                )
                await run.io_bound(memory.set_setting, "wiggle", "on" if e.value else "off")

            ui.switch(
                "Animación de la lechuza al pensar",
                value=memory.get_setting("wiggle", "on") == "on",
                on_change=_on_wiggle_change,
            )

        with ui.expansion("Carpetas habilitadas").classes("w-full"):
            ui.label("Carpetas por defecto (sin proyecto abierto) donde Lechu puede leer/escribir.") \
                .classes("text-caption text-gray-500")

            @ui.refreshable
            def folders_panel() -> None:
                for folder in CONFIG.filesystem.whitelisted_folders:
                    with ui.row().classes("items-center justify-between w-full"):
                        ui.label(str(folder)).classes("text-caption")
                        ui.button(
                            icon="delete",
                            on_click=lambda f=folder: (remove_whitelisted_folder(str(f)), folders_panel.refresh()),
                        ).props("flat dense round color=red")

            folders_panel()

            async def _add_folder() -> None:
                result = await app.native.main_window.create_file_dialog(dialog_type=webview.FileDialog.FOLDER)
                if not result:
                    return
                path = result[0] if isinstance(result, (list, tuple)) else result
                await run.io_bound(add_whitelisted_folder, path)
                folders_panel.refresh()

            ui.button("+ Agregar carpeta", on_click=_add_folder).classes("w-full")

        with ui.expansion("Memoria").classes("w-full"):
            @ui.refreshable
            def memory_panel() -> None:
                facts = memory.recall_facts()
                if facts:
                    rows = [{"category": f["category"], "key": f["key"], "value": f["value"]} for f in facts]
                    columns = [
                        {"name": "category", "label": "Categoría", "field": "category"},
                        {"name": "key", "label": "Clave", "field": "key"},
                        {"name": "value", "label": "Valor", "field": "value"},
                    ]
                    ui.table(rows=rows, columns=columns, row_key="key").classes("w-full")
                else:
                    ui.label("Todavía no hay hechos guardados.").classes("text-caption text-gray-500")

                cat_input = ui.input("Categoría")
                key_input = ui.input("Clave")
                val_input = ui.input("Valor")

                def _add_fact() -> None:
                    if cat_input.value and key_input.value and val_input.value:
                        memory.remember_fact(cat_input.value, key_input.value, val_input.value, source="user")
                        cat_input.value = ""
                        key_input.value = ""
                        val_input.value = ""
                        memory_panel.refresh()

                ui.button("Guardar", on_click=_add_fact)

                if facts:
                    del_options = {f["id"]: f"{f['category']}/{f['key']}" for f in facts}
                    del_select = ui.select(del_options, label="Borrar hecho")

                    def _delete_fact() -> None:
                        if del_select.value is not None:
                            memory.delete_fact(del_select.value)
                            memory_panel.refresh()

                    ui.button("Borrar hecho", on_click=_delete_fact)

            memory_panel()

        with ui.expansion("Skills").classes("w-full"):
            @ui.refreshable
            def skills_panel() -> None:
                for skill in state.skills:
                    ui.label(f"{skill.name} — {skill.description}").classes("text-sm")

                def _reload_skills() -> None:
                    state.skills = load_skills(SKILLS_DIR)
                    skills_panel.refresh()

                ui.button("Recargar skills", on_click=_reload_skills)

            skills_panel()

        with ui.expansion("Conexiones").classes("w-full"):
            ui.label("Clima").classes("font-bold text-sm")
            ui.label("Activo vía Open-Meteo — no requiere configuración.") \
                .classes("text-caption text-gray-500")

            ui.separator()
            ui.label("Ubicación").classes("font-bold text-sm")
            ui.label("Detección automática por IP, con GPS del sistema si está autorizado.") \
                .classes("text-caption text-gray-500")

            @ui.refreshable
            def location_panel() -> None:
                loc = location_service.get_cached_location()
                if loc:
                    place = ", ".join(p for p in (loc["city"], loc["region"], loc["country"]) if p)
                    tag = "exacta, GPS" if loc.get("source") == "gps" else "aproximada, IP"
                    ui.label(f"📍 {place} — {tag}").classes("text-caption")
                else:
                    ui.label("No se pudo detectar").classes("text-caption")

                async def _refresh_ip() -> None:
                    result = await run.io_bound(location_service.resolve_location)
                    location_panel.refresh()
                    ui.notify(
                        "✅ Ubicación actualizada" if result else "❌ No se pudo detectar la ubicación",
                        type="positive" if result else "negative",
                    )

                async def _refresh_gps() -> None:
                    result = await run.io_bound(location_service.resolve_location_gps)
                    location_panel.refresh()
                    ui.notify(
                        "✅ Ubicación exacta obtenida"
                        if result else "❌ No se pudo obtener GPS (¿permiso denegado o Servicios de Ubicación apagados?)",
                        type="positive" if result else "negative",
                    )

                with ui.row().classes("items-center"):
                    ui.button(
                        "Actualizar por IP", on_click=lambda: asyncio.create_task(_refresh_ip())
                    ).props("dense outline")
                    ui.button(
                        "Usar ubicación exacta (GPS)", on_click=lambda: asyncio.create_task(_refresh_gps())
                    ).props("dense outline")

            location_panel()

            ui.separator()
            ui.label("Maps").classes("font-bold text-sm")
            ui.label("Distancias y direcciones vía OpenRouteService.") \
                .classes("text-caption text-gray-500")

            @ui.refreshable
            def maps_panel() -> None:
                has_key = get_secret(maps_tools.SECRET_KEY) is not None
                ui.label("🔑 API key configurada" if has_key else "Sin configurar").classes("text-caption")

                key_input = ui.input(placeholder="Pegá tu API key de OpenRouteService...") \
                    .props("type=password dense").classes("w-full")

                def _save_key() -> None:
                    if key_input.value:
                        set_secret(maps_tools.SECRET_KEY, key_input.value)
                        key_input.value = ""
                        maps_panel.refresh()
                        ui.notify("API key guardada", type="positive")

                def _remove_key() -> None:
                    delete_secret(maps_tools.SECRET_KEY)
                    maps_panel.refresh()
                    ui.notify("API key eliminada", type="positive")

                async def _test_key() -> None:
                    ok = await run.io_bound(maps_tools.test_connection)
                    ui.notify(
                        "✅ Conexión OK" if ok else "❌ No se pudo validar la conexión (¿guardaste una key válida?)",
                        type="positive" if ok else "negative",
                    )

                with ui.row().classes("w-full items-center"):
                    ui.button("Guardar", on_click=_save_key).props("dense")
                    ui.button("Probar conexión", on_click=lambda: asyncio.create_task(_test_key())).props("dense outline")
                    if has_key:
                        ui.button("Quitar", on_click=_remove_key).props("dense outline color=red")

            maps_panel()

            ui.separator()
            ui.label("Búsqueda web").classes("font-bold text-sm")
            ui.label("Wikipedia activa vía su API pública — no requiere configuración.") \
                .classes("text-caption text-gray-500")
            ui.label("Google Custom Search requiere API key + Search Engine ID (cx).") \
                .classes("text-caption text-gray-500")
            ui.label(
                "Cada búsqueda prueba ~3 formas distintas de preguntar (3 consultas reales) - "
                "con el plan gratis de 100/día alcanza para ~33 búsquedas."
            ).classes("text-caption text-gray-500")

            @ui.refreshable
            def websearch_panel() -> None:
                has_key = (
                    get_secret(websearch_tools.API_KEY_SECRET) is not None
                    and get_secret(websearch_tools.CX_SECRET) is not None
                )
                ui.label("🔑 Configurado" if has_key else "Sin configurar").classes("text-caption")

                key_input = ui.input(placeholder="API key de Google Custom Search...") \
                    .props("type=password dense").classes("w-full")
                cx_input = ui.input(placeholder="Search Engine ID (cx)...") \
                    .props("dense").classes("w-full")

                def _save_key() -> None:
                    if key_input.value and cx_input.value:
                        set_secret(websearch_tools.API_KEY_SECRET, key_input.value)
                        set_secret(websearch_tools.CX_SECRET, cx_input.value)
                        key_input.value = ""
                        cx_input.value = ""
                        websearch_panel.refresh()
                        ui.notify("Búsqueda web guardada", type="positive")

                def _remove_key() -> None:
                    delete_secret(websearch_tools.API_KEY_SECRET)
                    delete_secret(websearch_tools.CX_SECRET)
                    websearch_panel.refresh()
                    ui.notify("Búsqueda web eliminada", type="positive")

                async def _test_key() -> None:
                    ok = await run.io_bound(websearch_tools.test_connection)
                    ui.notify(
                        "✅ Conexión OK" if ok else "❌ No se pudo validar la conexión (¿key/cx correctos?)",
                        type="positive" if ok else "negative",
                    )

                with ui.row().classes("w-full items-center"):
                    ui.button("Guardar", on_click=_save_key).props("dense")
                    ui.button("Probar conexión", on_click=lambda: asyncio.create_task(_test_key())).props("dense outline")
                    if has_key:
                        ui.button("Quitar", on_click=_remove_key).props("dense outline color=red")

            websearch_panel()

            ui.separator()
            ui.label("Google (Gmail, Drive, Calendar)").classes("font-bold text-sm")
            ui.label("Lectura y escritura con permiso, vía OAuth de Google.") \
                .classes("text-caption text-gray-500")

            @ui.refreshable
            def google_panel() -> None:
                connected_email = google_auth.get_connected_email()
                has_config = google_auth.has_client_config()

                if connected_email:
                    ui.label(f"✅ Conectado como {connected_email}").classes("text-caption")
                elif has_config:
                    ui.label("Credenciales guardadas, todavía no conectado.").classes("text-caption")
                else:
                    ui.label("Sin configurar").classes("text-caption")

                client_id_input = ui.input(placeholder="Client ID de Google Cloud") \
                    .props("dense").classes("w-full")
                client_secret_input = ui.input(placeholder="Client Secret de Google Cloud") \
                    .props("type=password dense").classes("w-full")

                def _save_client_config() -> None:
                    if client_id_input.value and client_secret_input.value:
                        google_auth.save_client_config(client_id_input.value, client_secret_input.value)
                        client_id_input.value = ""
                        client_secret_input.value = ""
                        google_panel.refresh()
                        ui.notify("Credenciales de Google guardadas", type="positive")

                async def _connect() -> None:
                    try:
                        email = await run.io_bound(google_auth.run_oauth_flow)
                    except Exception as e:
                        ui.notify(f"No se pudo conectar: {e}", type="negative")
                        return
                    google_panel.refresh()
                    ui.notify(f"Conectado como {email}", type="positive")

                def _disconnect() -> None:
                    google_auth.disconnect()
                    google_panel.refresh()
                    ui.notify("Desconectado de Google", type="positive")

                with ui.row().classes("w-full items-center"):
                    ui.button("Guardar credenciales", on_click=_save_client_config).props("dense")
                    connect_btn = ui.button(
                        "Conectar con Google", on_click=lambda: asyncio.create_task(_connect())
                    ).props("dense outline")
                    if not has_config:
                        connect_btn.props("disable")
                    if connected_email:
                        ui.button("Desconectar", on_click=_disconnect).props("dense outline color=red")

            google_panel()

        ui.separator()
        ui.label("Zona de peligro").classes("font-bold text-sm text-red")

        async def _reset_facts() -> None:
            if await _confirm_dialog("¿Borrar TODOS los hechos guardados? Esta acción no se puede deshacer."):
                await run.io_bound(memory.delete_all_facts)
                memory_panel.refresh()
                ui.notify("Hechos borrados", type="positive")

        async def _reset_conversations() -> None:
            if await _confirm_dialog("¿Borrar TODAS las conversaciones? Esta acción no se puede deshacer."):
                await run.io_bound(memory.delete_all_conversations)
                refs = refs_holder["refs"]
                await start_new_conversation(state, _effective_agent(state))
                render_chat_history(state, refs.chat_container)
                render_canvas(state, refs)
                refs_holder["refresh_history"]()
                ui.notify("Conversaciones borradas", type="positive")

        ui.button("Borrar todos los hechos", on_click=_reset_facts).props("outline color=red").classes("w-full")
        ui.button("Borrar todas las conversaciones", on_click=_reset_conversations).props("outline color=red").classes("w-full")

    dialog.open()


async def _prime_location() -> None:
    await run.io_bound(location_service.resolve_location)
    await run.io_bound(location_service.resolve_location_gps)


def main() -> None:
    build_page()
    app.on_startup(_prime_location)
    app.on_shutdown(get_client().aclose)
    ui.run(
        native=True,
        window_size=(1400, 900),
        title="Lechu",
        favicon=str(Path(__file__).resolve().parent / "assets" / "lechu_mark.svg"),
        reload=False,
        show=True,
    )


if __name__ in {"__main__", "__mp_main__"}:
    main()
